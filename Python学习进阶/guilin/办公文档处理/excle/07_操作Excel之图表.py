# !/usr/bin/python3
# _*_ coding:utf-8 _*_
#
# Copyright (C) 2023 - 2023 pgl, Inc. All Rights Reserved
#
# @Time    : 2023/2/22 9:14
# @Author  : pgl
# @File    : 07_操作Excel之图表.py
# @IDE     : PyCharm

from openpyxl import Workbook
from datetime import date  # 时间日期
from openpyxl.chart import LineChart, Reference, PieChart, BarChart  # 创建图表


def set_img1():
    wb = Workbook();
    sh = wb.active

    rows = [
        ['date', 'group1', 'group2', 'group3'],
        [date(2050, 12, 1), 40, 30, 25],
        [date(2050, 12, 2), 40, 28, 25],
        [date(2050, 12, 3), 20, 30, 75],
        [date(2050, 12, 4), 10, 42, 54],
        [date(2050, 12, 5), 50, 57, 57],
        [date(2050, 12, 6), 60, 25, 87],
        [date(2050, 12, 7), 27, 42, 35]
    ]

    for row in rows:
        sh.append(row)

    # 绘图
    # 折线图
    lc = LineChart()
    lc.style = 10  # 图标线的粗细
    lc.title = 'Line Chart'  # 图标名字
    # x y轴 的信息
    lc.x_axis.title = 'test_number'  # x轴统计具体数字
    lc.y_axis.title = 'size'  # y轴统计大小

    # 填充数据
    # min_col最小列索引          min_row 最小行索引
    # max_col最大列索引          max_row最大行索引
    data = Reference(sh, min_col=2, max_col=4, min_row=1, max_row=8)  # 告诉我们范围怎样选取
    lc.add_data(data, titles_from_data=True)

    # 添加图表 增加到Excel文件
    sh.add_chart(lc, 'A9')

    wb.save('res/图表Excel.xlsx')


def set_img2():
    wb = Workbook();
    sh = wb.active

    data = [
        ['名称', '数值', ],
        ['苹果', 40],
        ['草莓', 74],
        ['叶子', 65],
        ['荔枝', 50],
        ['葡萄', 25],
    ]

    for row in data:
        sh.append(row)

    # 绘图
    # 饼图
    pie = PieChart()
    lables = Reference(sh, min_col=1, min_row=2, max_row=6)
    data = Reference(sh, min_col=2, min_row=1, max_row=6)

    # 加数据
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(lables)

    # 添加图表 增加到Excel文件
    sh.add_chart(pie, 'A8')

    wb.save('res/图表Excel1.xlsx')


def set_img3():
    wb = Workbook();
    sh = wb.active

    data = [
        ['数值', '组1', '组2'],
        [1, 10, 25],
        [2, 23, 24],
        [3, 12, 25],
        [4, 12, 3],
        [5, 32, 12],
        [6, 25, 24],
    ]

    for row in data:
        sh.append(row)

    #     绘画图表
    bar = BarChart()
    data = Reference(sh, min_col=2, min_row=1, max_row=7, max_col=3)
    cats = Reference(sh, min_col=1, min_row=2, max_row=7)

    # 写入图表信息
    bar.add_data(data, titles_from_data=True)
    # 写入分类信息
    bar.set_categories(cats)
    # 图表加到第几个单元格
    sh.add_chart(bar, 'A7')

    wb.save('res/图表Excel2.xlsx')


if __name__ == '__main__':
    # set_img1()
    set_img2()
    # set_img3()
