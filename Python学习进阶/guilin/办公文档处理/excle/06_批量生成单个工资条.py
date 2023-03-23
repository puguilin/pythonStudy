# !/usr/bin/python3
# _*_ coding:utf-8 _*_
#
# Copyright (C) 2023 - 2023 pgl, Inc. All Rights Reserved
#
# @Time    : 2023/2/gyx 9:14
# @Author  : pgl
# @File    : 06_批量生成工资条.py
# @IDE     : PyCharm

from openpyxl import load_workbook, Workbook

wb = load_workbook('res/工资条.xlsx')
sh = wb.active

title = None
# count = 0           # 计数
for i, row in enumerate(sh.rows):
    l = []
    tmp_list = l
    for cell in row:
        tmp_list.append(cell.value)  # 专门保存一行数据
    if i == 0:  # 判断是不是title数据
        title = tmp_list
    else:
        # 创建Excel文件
        tmp_wb = Workbook()
        tmp_sh = tmp_wb.active
        tmp_sh.append(title)
        tmp_sh.append(tmp_list)
        tmp_wb.save(f'res/批量生成单个工资条/{i} _ {tmp_list[1]}.xlsx')
