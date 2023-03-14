# !/usr/bin/python3
# _*_ coding:utf-8 _*_
#
# Copyright (C) 2023 - 2023 pgl, Inc. All Rights Reserved
#
# @Time    : 2023/2/22 9:14
# @Author  : pgl
# @File    : 09_美化Excel.py
# @IDE     : PyCharm
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment, PatternFill  # Excel美化

wb = Workbook()
sh = wb.active
# font =  Font(name = '微软雅黑' , size = '30' , bold= 'True' , italic=True ,color = colors.BLUE)
font = Font(name='微软雅黑', size='30', bold='True', italic=True, color='FF0000')
# 加粗           斜体          颜色
sh['B2'] = 'Hello'
sh['B2'].font = font

# 修改单元格
sh.row_dimensions[6].height = 30  # 高度
sh.column_dimensions['C'].width = 15  # 宽度
sh['C6'] = 'Python'

# 文字对齐
sh['C6'].alignment = Alignment(horizontal='right', vertical='top')  # 上下左右对齐

# 单元格填充
sh['H5'].fill = PatternFill('solid', '00FF00')

wb.save('res/美化Excel.xlsx')
