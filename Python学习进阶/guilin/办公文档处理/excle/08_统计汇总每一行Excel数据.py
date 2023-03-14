# !/usr/bin/python3
# _*_ coding:utf-8 _*_
#
# Copyright (C) 2023 - 2023 pgl, Inc. All Rights Reserved
#
# @Time    : 2023/2/22 9:14
# @Author  : pgl
# @File    : 08_统计汇总每一行Excel数据.py
# @IDE     : PyCharm

# 导包
from openpyxl import load_workbook

# 打开文件
wb = load_workbook('res/abc.xlsx')
# 激活工作博
sh = wb.active

# 读取数据
all = []
for row in sh.iter_rows(min_row=2, min_col=2):  # min_row=2第二行开始   min_col=2 第二列开始
    tmp_list = []
    for cell in row:
        tmp_list.append(cell.value)
    all.append(tmp_list)

# 汇总
all_count = []  # 记录每一行的数据和
for a in all:
    count = 0  # 记录单行数据和
    for n in a:
        tmp = str(n)  # 转换成str，是为了获取判断是否是数字
        if tmp.isdigit():  # 判断是否为数字
            count += n
    all_count.append(count)
    # print(count)
# print(all)

print(sh.max_row)  # 获取最高行
print(sh.max_column)  # 获取最高列

max = sh.max_column + 1  # 记录下最高列是多少
sh.cell(1, max).value = '汇总'

row_num = 2
for n in all_count:  # 获取所有的统计数据
    sh.cell(row_num, max).value = n
    row_num += 1

# 保存
wb.save(r'res/abc123.xlsx')


