# !/usr/bin/python3
# _*_ coding:utf-8 _*_
#
# Copyright (C) 2023 - 2023 pgl, Inc. All Rights Reserved
#
# @Time    : 2023/2/gyx 9:14
# @Author  : pgl
# @File    : 05_快速查找重复数据.py
# @IDE     : PyCharm

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook('res/加班时间.xlsx')
sh = wb.active  # 激活工作表

index = []  # 记录哪一行数据是重复的
tmp = []  # 记录没有重复的数据
for i, cell in enumerate(sh['B']):  # 根据某一列判断是否重复
    # print(cell.value)
    flag = cell.value not in tmp
    if flag:
        tmp.append(cell.value)
    else:
        index.append(i)

# print(tmp)
# print(index)
fill = PatternFill('solid', 'AEEEEE')  # solid 实心 AEEEEE 颜色

for i, row in enumerate(sh.rows):  # 填充
    if i in index:
        for cell in row:
            cell.fill = fill
        print(f"第{i}行是重复数据")

wb.save('res/查找重复数据.xlsx')
