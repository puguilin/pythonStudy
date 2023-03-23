# !/usr/bin/python3
# _*_ coding:utf-8 _*_
#
# Copyright (C) 2023 - 2023 pgl, Inc. All Rights Reserved
#
# @Time    : 2023/2/gyx 9:14
# @Author  : pgl
# @File    : 04_python拆分Excel.py
# @IDE     : PyCharm

from openpyxl import load_workbook


def chaifen():
    # 拆分表格 按照列的值大小拆分
    wb = load_workbook(r'res/a1.xlsx')
    # 激活工作簿
    sh = wb.active
    # 少于300的数据
    data1 = []
    # 大于300的数据
    data2 = []

    # 读数据
    for row in sh.rows:  # 获取列
        num = row[2].value # 获取列值
        # print(num)
        if num >= 300:
            data2.append(row)
        else:
            data1.append(row)

    data1_sh = wb.create_sheet('少于300')
    data2_sh = wb.create_sheet('大于300')

    for d in data1:
        t_list = []
        for t in d:
            t_list.append(t.value)
        data1_sh.append(t_list)

    for d in data2:
        t_list = []
        for t in d:
            t_list.append(t.value)
        data2_sh.append(t_list)
    # 保存
    wb.save(r'res\test4.xlsx')
    print('结束')


if __name__ == '__main__':
    chaifen()
