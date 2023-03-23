# !/usr/bin/python3
# _*_ coding:utf-8 _*_
#
# Copyright (C) 2023 - 2023 pgl, Inc. All Rights Reserved
#
# @Time    : 2023/2/gyx 9:14
# @Author  : pgl
# @File    : 03_合并Excel文件.py
# @IDE     : PyCharm

# 操作Excel
from openpyxl import load_workbook, Workbook

# 打开Excel
wb = load_workbook('res/a1.xlsx')
# 激活多个工作表

all = []  # 存储所有数据

sh1 = wb['Sheet1']
sh2 = wb['Sheet2']
# 读取数据
# 合并几个就读几个 第一个
for row in sh1.rows:  # 获取所有行
    tmp_list = []  # 用来存储一行的值
    for cell in row:  # 获取所有列
        tmp_list.append(cell.value)  # 获取单元格的值
    all.append(tmp_list)  # 整合所有数据

    # 合并几个就读几个 第二个
for row in sh2.rows:  # 获取所有行python拆分Excel.py
    tmp_list = []  # 用来存储一行的值
    for cell in row:  # 获取所有列
        tmp_list.append(cell.value)  # 获取单元格的值
    all.append(tmp_list)  # 整合所有数据

# 保存所有数据
wb = Workbook()
sh = wb.active
for row in all:  # 遍历所有数据
    sh.append(row)

wb.save('./res/test5.xlsx')

