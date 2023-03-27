# -*- codeing = utf-8 -*-
# @Time : 2022/3/25 23:45
# @Author : 邢继森
# @File : 创建Excel.py
# Software : PyCharm
from openpyxl import Workbook


def new():
    wb = Workbook()
    # 激活默认表
    sh1 = wb.active
    # 创建工作簿和文件
    sh2 = wb.create_sheet('电视信息')  # 工作薄信息
    wb.save('res/test2.xlsx')  # 文件名称

    # 填充数据


def set_value():
    from openpyxl import Workbook
    wb = Workbook()
    # 激活默认表
    sh = wb.active
    # 写入数据

    for i in range(1, 10):
        sh[f'A{i}'] = 'Hello'
        sh[f'B{i}'] = '你好'
        sh[f'C{i}'] = '123456'
        wb.save('res/退单审核.xlsx')


# 批量写入数据
def set_many_value():
    from openpyxl import Workbook
    wb = Workbook()
    # 激活默认表
    sh = wb.active

    rows = [
        ['沙雕', 123],
        ['家富煌钢', 456],
        ['构时', 789],
        ['代风貌', 000]
    ]
    for row in rows:
        sh.append(row)
    wb.save('test3.xlsx')


# 创建带有标题的excel

def biaotou():
    wb1 = Workbook()
    # 创建工作簿
    sh3 = wb1.create_sheet('首页数据123')  # 工作薄信息
    sh2 = wb1.create_sheet('首页数据')  # 工作薄信息
    # sh2 = wb1['首页数据']
    one = [
        '店铺名称',
        '评分',
        '评论',
        '地址',
        '人均价格'
    ]
    two = [
        '店铺名称02',
        '评分02',
        '评论02',
        '地址02',
        '人均价格02'
    ]
    sh3.append(one)
    sh2.append(two)
    wb1.save('res/test2.xlsx')  # 保存文件


if __name__ == '__main__':
    biaotou()
    # new()
    # s et_value()
    # set_many_value()
