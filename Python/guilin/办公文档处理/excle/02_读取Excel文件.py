# !/usr/bin/python3
# _*_ coding:utf-8 _*_
#
# Copyright (C) 2023 - 2023 pgl, Inc. All Rights Reserved 
#
# @Time    : 2023/2/gyx 9:14
# @Author  : pgl
# @File    : 02_读取Excel文件.py.py
# @IDE     : PyCharm


from openpyxl import load_workbook
from openpyxl import Workbook

workbook = load_workbook('res/学生明细表.xlsx')
print(workbook.sheetnames)
sheet = workbook[workbook.sheetnames[0]]
print(sheet.title)
for row in range(2, 7):
    for col in range(65, 70):
        cell_index = chr(col) + str(row)
        print(sheet[cell_index].value, end='\t')
    print()
